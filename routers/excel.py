import io
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, Form, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse

from dependencies.auth import get_current_user
from services.plate_utils import normalize_plate_value
from services.excel_utils import apply_excel_style, workbook_to_bytes

router = APIRouter(
    prefix="/api",
    tags=["excel"],
    dependencies=[Depends(get_current_user)],
)

_EXPORT_HEADERS = [
    "#", "رقم اللوحة", "نوع المركبة", "الشارع",
    "تفاصيل الموقع", "ملاحظات", "اسم المسجّل",
    "تاريخ التسجيل", "GPS",
]
_COL_WIDTHS = [5, 20, 14, 22, 25, 20, 16, 18, 22]


def _clean_sheet_name(name: str) -> str:
    for ch in r'/\?*[]':
        name = name.replace(ch, "")
    return (name or "بيانات المركبات")[:31]


@router.post("/export-excel")
async def export_excel(
    rows_json:  str = Form("[]"),
    sheet_name: str = Form("بيانات المركبات"),
):
    sheet_name = _clean_sheet_name(sheet_name.strip())

    try:
        rows = json.loads(rows_json)
    except Exception:
        raise HTTPException(status_code=400, detail="تنسيق JSON خاطئ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True

    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hfill = PatternFill("solid", start_color="1F4E79")
    ha    = Alignment(horizontal="center", vertical="center")
    ca    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df    = Font(name="Arial", size=11)
    thin  = Side(style="thin", color="BFBFBF")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    fe    = PatternFill("solid", start_color="D6E4F0")
    fo    = PatternFill("solid", start_color="FFFFFF")

    for col, h in enumerate(_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = ha; cell.border = brd
    ws.row_dimensions[1].height = 30

    # Filter invalid plates
    valid_rows = []
    for r in rows:
        normalized, ok = normalize_plate_value(full_raw=r.get("full_plate", ""))
        if not ok:
            continue
        rr = dict(r)
        rr["full_plate"] = normalized
        valid_rows.append(rr)

    for i, r in enumerate(valid_rows, 1):
        fill = fe if i % 2 == 0 else fo
        vals = [
            i,
            r.get("full_plate", ""),
            r.get("vehicle_type", "ملاكى"),
            r.get("street_name", "غير محدد"),
            r.get("location_details", ""),
            r.get("notes", ""),
            r.get("recorder_name", ""),
            r.get("recording_date", ""),
            r.get("gps", ""),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=v)
            cell.font = df; cell.alignment = ca
            cell.border = brd; cell.fill = fill

    for col, w in zip("ABCDEFGHI", _COL_WIDTHS):
        ws.column_dimensions[col].width = w

    content = workbook_to_bytes(wb)
    filename = f"تفريغ_{sheet_name}.xlsx"

    return StreamingResponse(
        io.BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/export-field-check")
async def export_field_check(
    rows_json:  str = Form("[]"),
    sheet_name: str = Form("التشيك الميداني"),
):
    """Same as export-excel but with a different default sheet name."""
    sheet_name = _clean_sheet_name(sheet_name.strip() or "التشيك الميداني")

    try:
        rows = json.loads(rows_json)
    except Exception:
        raise HTTPException(status_code=400, detail="تنسيق JSON خاطئ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True

    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hfill = PatternFill("solid", start_color="0D6B5E")   # teal for field-check
    ha    = Alignment(horizontal="center", vertical="center")
    ca    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df    = Font(name="Arial", size=11)
    thin  = Side(style="thin", color="BFBFBF")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    fe    = PatternFill("solid", start_color="E0F2F1")
    fo    = PatternFill("solid", start_color="FFFFFF")

    headers = ["#", "رقم اللوحة", "نوع المركبة", "الشارع",
               "تفاصيل الموقع", "ملاحظات", "اسم المسجّل",
               "تاريخ التسجيل", "GPS"]
    col_widths = [5, 20, 14, 22, 25, 20, 16, 18, 22]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = ha; cell.border = brd
    ws.row_dimensions[1].height = 30

    valid_rows = []
    for r in rows:
        normalized, ok = normalize_plate_value(full_raw=r.get("full_plate", ""))
        if not ok:
            continue
        rr = dict(r)
        rr["full_plate"] = normalized
        valid_rows.append(rr)

    for i, r in enumerate(valid_rows, 1):
        fill = fe if i % 2 == 0 else fo
        vals = [
            i,
            r.get("full_plate", ""),
            r.get("vehicle_type", "ملاكى"),
            r.get("street_name", "غير محدد"),
            r.get("location_details", ""),
            r.get("notes", ""),
            r.get("recorder_name", ""),
            r.get("recording_date", ""),
            r.get("gps", ""),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=v)
            cell.font = df; cell.alignment = ca
            cell.border = brd; cell.fill = fill

    for col, w in zip("ABCDEFGHI", col_widths):
        ws.column_dimensions[col].width = w

    content  = workbook_to_bytes(wb)
    filename = f"اللوحات_المطابقة_{sheet_name}.xlsx"

    return StreamingResponse(
        io.BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    if not file:
        raise HTTPException(status_code=400, detail="لم يتم رفع ملف")

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        ws = wb.active
        rows_out = []
        headers  = []

        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri == 0:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            if all(c is None for c in row):
                continue

            def col(name, fallback):
                for i, h in enumerate(headers):
                    if name in h:
                        return (
                            str(row[i]).strip()
                            if i < len(row) and row[i] is not None
                            else ""
                        )
                return (
                    str(row[fallback]).strip()
                    if fallback < len(row) and row[fallback] is not None
                    else ""
                )

            rows_out.append({
                "full_plate":       col("اللوحة", 1),
                "vehicle_type":     col("المركبة", 2),
                "street_name":      col("الشارع", 3),
                "location_details": col("الموقع", 4),
                "notes":            col("ملاحظات", 5),
                "recorder_name":    col("المسجّل", 6),
                "recording_date":   col("التسجيل", 7),
                "gps":              col("GPS", 8),
            })

        return JSONResponse({"rows": rows_out, "total": len(rows_out)})

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))